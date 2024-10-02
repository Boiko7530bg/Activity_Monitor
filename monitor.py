import tkinter as tk
import threading
import pystray
from PIL import Image, ImageDraw
import os
import getpass
import time
from datetime import datetime, timedelta
import pygetwindow as gw
from openpyxl import Workbook, load_workbook
from pynput import mouse, keyboard
from threading import Lock
import sys  # Import sys to check if running as executable

last_activity_time = datetime.now()
last_window_title = None
idle_start_time = None
current_activity_start_time = datetime.now()
current_date = datetime.now().date()
last_backup_time = datetime.now()

idle_threshold = 60  # Start idling after - (seconds)

backup_interval_in_hours = 0.50  # Backup interval in hours (0.50 = 30mins)
backup_location = os.path.join(os.path.expanduser("~"), "Documents", "Activity_Backup")

documents_folder = os.path.join(os.path.expanduser("~"), "Documents")
activity_folder = os.path.join(documents_folder, "Activity")

excel_lock = Lock()

sleep_interval = 1  # Sleep interval in seconds

total_idle_time = timedelta()
total_working_time = timedelta()
total_idle_time_lock = Lock()
total_working_time_lock = Lock()

icon = None


def create_image():
    width = 64
    height = 64
    color1 = 'blue'
    color2 = 'white'

    image = Image.new('RGB', (width, height), color1)
    draw = ImageDraw.Draw(image)
    draw.rectangle(
        [(width // 2, 0), (width, height // 2)],
        fill=color2)
    draw.rectangle(
        [(0, height // 2), (width // 2, height)],
        fill=color2)
    return image


def on_quit(icon_param, item):
    icon_param.visible = False
    root.quit()
def show_window(icon_param, item):
    icon_param.visible = False
    root.after(0, root.deiconify)


def withdraw_window():
    root.withdraw()
    icon.visible = True


def on_minimize(event):
    if root.state() == 'iconic':
        withdraw_window()


def get_log_path():
    file_name = f"{getpass.getuser()}_{current_date.strftime('%Y-%m-%d')}.xlsx"
    return os.path.join(activity_folder, file_name)


def get_backup_log_path():
    log_path = get_log_path()
    base_name, ext = os.path.splitext(os.path.basename(log_path))
    backup_file_name = f"{base_name}_backup{ext}"
    return os.path.join(backup_location, backup_file_name)


def ensure_activity_folder():
    if not os.path.exists(activity_folder):
        os.makedirs(activity_folder)


def ensure_backup_folder():
    if not os.path.exists(backup_location):
        os.makedirs(backup_location)


def initialize_workbook():
    ensure_activity_folder()

    log_path = get_log_path()

    if not os.path.exists(log_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Activity Log"

        headers = ["LoginName", "MOTION_APPLICATION_CR", "MOTION_TYPE_CR", "START_TIME_DT", "END_TIME_DT", "TotalTime"]
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num).value = header

        wb.create_sheet(title="Summary")
        wb["Summary"].cell(row=1, column=1).value = "Idle Hours"
        wb.save(log_path)
    return load_workbook(log_path)


def update_activity():
    global last_activity_time
    last_activity_time = datetime.now()


def get_active_window():
    try:
        active_window = gw.getActiveWindow()
        if active_window:
            window_title = active_window.title
            return f"{window_title}"
    except Exception as e:
        print(f"Error getting active window: {e}")
        return "Unknown Window"
    return "Unknown Window"


def log_to_excel(activity_type, window_title, start_time, end_time):
    with excel_lock:
        wb = initialize_workbook()
        ws = wb.active

        next_row = ws.max_row + 1

        # Populate columns based on format
        ws.cell(row=next_row, column=1).value = getpass.getuser()
        ws.cell(row=next_row, column=2).value = window_title
        ws.cell(row=next_row, column=3).value = activity_type
        ws.cell(row=next_row, column=4).value = start_time.strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=next_row, column=5).value = end_time.strftime('%Y-%m-%d %H:%M:%S')

        total_time = end_time - start_time
        total_seconds = int(total_time.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        formatted_time = f"{hours}:{minutes:02}:{seconds:02}"  # h:mm:ss format
        ws.cell(row=next_row, column=6).value = formatted_time

        # Try to save the workbook, handling the case where the file might be locked
        max_retries = 5
        retries = 0
        while retries < max_retries:
            try:
                wb.save(get_log_path())
                break
            except PermissionError:
                retries += 1
                print(f"Unable to save {get_log_path()}. It might be open. Retrying in 5 seconds...")
                time.sleep(5)
        else:
            print(f"Failed to save {get_log_path()} after {max_retries} attempts.")


def on_click(x, y, button, pressed):
    if pressed:
        update_activity()


def on_move(x, y):
    update_activity()


def on_press(key):
    update_activity()


def backup_excel_file():
    ensure_backup_folder()
    try:
        log_path = get_log_path()
        backup_file_path = get_backup_log_path()
        import shutil
        shutil.copyfile(log_path, backup_file_path)
        print(f"Backup created at {backup_file_path}")
    except Exception as e:
        print(f"Failed to create backup: {e}")


def monitor_activity():
    global last_activity_time, idle_start_time, current_activity_start_time, last_window_title, current_date, last_backup_time
    global total_idle_time, total_working_time, is_idle  # Added is_idle to global

    is_idle = False

    while True:
        current_time = datetime.now()
        time_since_last_activity = (current_time - last_activity_time).total_seconds()

        if time_since_last_activity >= idle_threshold:
            if not is_idle:
                is_idle = True
                idle_start_time = last_activity_time + timedelta(seconds=idle_threshold)
                if last_window_title and current_activity_start_time:
                    working_duration = idle_start_time - current_activity_start_time
                    with total_working_time_lock:
                        total_working_time += working_duration
                    log_to_excel("Working", last_window_title, current_activity_start_time, idle_start_time)
                current_activity_start_time = None
                last_window_title = None
        else:
            if is_idle:
                is_idle = False
                idle_end_time = last_activity_time
                idle_duration = idle_end_time - idle_start_time
                with total_idle_time_lock:
                    total_idle_time += idle_duration
                log_to_excel("Idle", "Idle Hours", idle_start_time, idle_end_time)
                idle_start_time = None
                current_activity_start_time = last_activity_time
                last_window_title = get_active_window()
            else:
                active_window = get_active_window()
                if active_window != last_window_title:
                    if last_window_title and current_activity_start_time:
                        duration = current_time - current_activity_start_time
                        with total_working_time_lock:
                            total_working_time += duration
                        log_to_excel("Working", last_window_title, current_activity_start_time, current_time)
                    current_activity_start_time = current_time
                    last_window_title = active_window

        if (current_time - last_backup_time).total_seconds() >= backup_interval_in_hours * 3600:
            backup_excel_file()
            last_backup_time = current_time

        time.sleep(sleep_interval)


def update_gui():
    current_time = datetime.now()

    with total_idle_time_lock:
        idle_time = total_idle_time
    with total_working_time_lock:
        working_time = total_working_time

    if is_idle:
        if idle_start_time:
            idle_time += current_time - idle_start_time
    else:
        if current_activity_start_time:
            working_time += current_time - current_activity_start_time

    total_time = idle_time + working_time

    idle_seconds = int(idle_time.total_seconds())
    working_seconds = int(working_time.total_seconds())
    total_seconds = int(total_time.total_seconds())

    total_hours, total_remainder = divmod(total_seconds, 3600)
    total_minutes, total_secs = divmod(total_remainder, 60)

    working_hours, working_remainder = divmod(working_seconds, 3600)
    working_minutes, working_secs = divmod(working_remainder, 60)

    idle_hours, idle_remainder = divmod(idle_seconds, 3600)
    idle_minutes, idle_secs = divmod(idle_remainder, 60)

    total_time_str = f"{total_hours}h {total_minutes}m {total_secs}s"
    working_time_str = f"{working_hours}h {working_minutes}m {working_secs}s"
    idle_time_str = f"{idle_hours}h {idle_minutes}m {idle_secs}s"

    # Update the labels
    total_label.config(text=f"Total Time: {total_time_str}")
    working_label.config(text=f"Working Time: {working_time_str}")
    idle_label.config(text=f"Idle Time: {idle_time_str}")

    root.after(1000, update_gui)  # Update every second

def setup_icon():
    global icon
    image = create_image()
    menu = pystray.Menu(
        pystray.MenuItem('Show', show_window),
        pystray.MenuItem('Exit', on_quit)
    )
    icon = pystray.Icon("name", image, "Activity Monitor", menu)
    icon.visible = False  # Initially invisible
    threading.Thread(target=icon.run, daemon=True).start()

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Activity Monitor")
    root.geometry("200x200")

    # Bind the minimize event
    root.bind("<Unmap>", on_minimize)

    # Handle the window close button
    root.protocol('WM_DELETE_WINDOW', root.iconify)

    # Create labels to display the times
    total_label = tk.Label(root, text="Total Time: 0h 0m 0s", font=("Helvetica", 12))
    total_label.pack(pady=5)

    working_label = tk.Label(root, text="Working Time: 0h 0m 0s", font=("Helvetica", 12))
    working_label.pack(pady=5)

    idle_label = tk.Label(root, text="Idle Time: 0h 0m 0s", font=("Helvetica", 12))
    idle_label.pack(pady=5)

    initialize_workbook()

    activity_thread = threading.Thread(target=monitor_activity, daemon=True)
    activity_thread.start()

    mouse_listener = mouse.Listener(on_click=on_click, on_move=on_move)
    keyboard_listener = keyboard.Listener(on_press=on_press)

    mouse_listener.start()
    keyboard_listener.start()

    update_gui()

    setup_icon()

    if getattr(sys, 'frozen', False):
        # If running as an executable, start minimized in system tray
        withdraw_window()
    else:
        # If running as script, show the main window
        root.deiconify()

    root.mainloop()

    mouse_listener.stop()
    keyboard_listener.stop()
